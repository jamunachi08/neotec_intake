"""Deterministic extraction for Excel (.xlsx) and CSV/TSV — no AI needed."""
import csv
import io


def _clean(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def _locate_header(grid: list[list]):
    """Pick the row with the most non-empty, mostly-text cells as the header."""
    best_i, best_score = 0, -1
    for i, row in enumerate(grid[:30]):
        cells = [c for c in row if str(_clean(c)) != ""]
        if len(cells) < 2:
            continue
        texty = sum(1 for c in cells if isinstance(c, str) and any(ch.isalpha() for ch in c))
        score = len(cells) + texty
        if score > best_score:
            best_i, best_score = i, score
    return best_i


def extract_csv(content: bytes, header_row: int = 0) -> dict:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delim = "\t" if sample.count("\t") > sample.count(",") else ","
    grid = [row for row in csv.reader(io.StringIO(text), delimiter=delim)]
    return _grid_to_payload(grid, header_row)


def extract_xlsx(content: bytes, sheet: str = "", header_row: int = 0) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    grid = [[_clean(c) for c in row] for row in ws.iter_rows(values_only=True)]
    return _grid_to_payload(grid, header_row)


def _grid_to_payload(grid: list[list], header_row: int = 0) -> dict:
    grid = [r for r in grid if any(str(_clean(c)) != "" for c in r)]
    if not grid:
        return {"fields": {}, "rows": [], "raw_text": ""}
    hidx = (header_row - 1) if header_row and header_row > 0 else _locate_header(grid)
    hidx = max(0, min(hidx, len(grid) - 1))
    headers = []
    seen = {}
    for j, h in enumerate(grid[hidx]):
        name = str(_clean(h)) or f"col_{j+1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    rows = []
    for r in grid[hidx + 1:]:
        rec = {}
        for j, h in enumerate(headers):
            rec[h] = _clean(r[j]) if j < len(r) else ""
        if any(str(v) != "" for v in rec.values()):
            rows.append(rec)
    return {"fields": {}, "rows": rows, "raw_text": "", "headers": headers}
